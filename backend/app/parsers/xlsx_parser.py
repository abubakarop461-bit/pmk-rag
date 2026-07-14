import openpyxl
from typing import List
from langchain_core.documents import Document
from app.parsers.base_parser import BaseParser
from loguru import logger

class XlsxParser(BaseParser):
    def parse(self, file_path: str) -> List[Document]:
        """
        Loads Excel workbook using openpyxl and extracts cell grid text page-by-page.
        """
        logger.info(f"XlsxParser starting extraction on: {file_path}")
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            documents = []
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                rows_text = []
                for row in sheet.iter_rows(values_only=True):
                    # Only include row if at least one cell has a value
                    if any(cell is not None for cell in row):
                        row_str = " | ".join(str(cell) if cell is not None else "" for cell in row)
                        rows_text.append(row_str)
                        
                sheet_content = f"--- Sheet: {sheet_name} ---\n" + "\n".join(rows_text)
                documents.append(
                    Document(
                        page_content=sheet_content.strip(),
                        metadata={
                            "source": file_path,
                            "sheet_name": sheet_name,
                            "page": len(documents) + 1
                        }
                    )
                )
            wb.close()
            logger.info(f"XlsxParser completed. Extracted {len(documents)} sheets from {file_path}.")
            return documents
        except Exception as e:
            logger.error(f"openpyxl failed to parse Excel workbook {file_path}: {e}")
            raise RuntimeError(f"Excel parsing failed: {e}")
